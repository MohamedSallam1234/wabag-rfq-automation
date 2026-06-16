"""Tests for rendering an RFQGeneration into a fresh .xlsx datasheet."""

import io

from openpyxl import load_workbook

from app.schemas.rfq import RFQField, RFQFieldConflict, RFQGeneration, RFQSection
from app.services.rfq.renderer import render_xlsx


def _generation() -> RFQGeneration:
    return RFQGeneration(
        equipment_tag="B-100",
        equipment_category="Blower",
        header={"Project": "Kohafa WWTP", "Client": "WABAG"},
        sections=[
            RFQSection(
                title="Process Data",
                fields=[
                    RFQField(
                        field="Capacity",
                        value=860,
                        unit="m3/hr",
                        confidence=0.9,
                        source_document="04_Hydraulic Profile",
                        source_location="Sheet Design, row Flow",
                        evidence="design flow 860 m3/hr",
                        status="extracted",
                    ),
                    RFQField(field="Head", value=None, status="tbd"),
                    RFQField(
                        field="Material",
                        status="conflict",
                        conflicts=[
                            RFQFieldConflict(
                                value="SS304",
                                source_document="01_Employer Spec",
                                source_location="Sec 3.2",
                                evidence="pump body SS304",
                            ),
                            RFQFieldConflict(
                                value="SS316",
                                source_document="02_Process",
                                source_location="Table 4",
                                evidence="material SS316",
                            ),
                        ],
                    ),
                    RFQField(field="Coupling guard", status="vtf"),
                    RFQField(field="Anchor bolts", value="By Vendor", status="vtf"),
                    RFQField(
                        field="Motor power",
                        status="vtf",
                        conflicts=[
                            RFQFieldConflict(
                                value="200 kW",
                                source_document="03_Equipment List",
                                source_location="row Motor",
                                evidence="motor 200 kW",
                            ),
                        ],
                    ),
                ],
            )
        ],
    )


def _flat(data: bytes) -> list[str]:
    workbook = load_workbook(io.BytesIO(data))
    sheet = workbook.active
    return [str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value is not None]


def test_render_xlsx_contains_header_section_and_columns() -> None:
    flat = _flat(render_xlsx(_generation()))

    assert "Kohafa WWTP" in flat  # header value
    assert "Process Data" in flat  # section title
    assert {
        "Field",
        "Value",
        "Unit",
        "Confidence",
        "Source Document",
        "Location",
        "Evidence",
        "Status",
    } <= set(flat)


def test_render_xlsx_extracted_field_shows_structured_provenance() -> None:
    flat = _flat(render_xlsx(_generation()))
    assert "860" in flat
    assert "04_Hydraulic Profile" in flat  # source document column
    assert "Sheet Design, row Flow" in flat  # location column
    assert "design flow 860 m3/hr" in flat  # evidence column


def test_render_xlsx_conflict_fills_both_candidates_positionally() -> None:
    flat = _flat(render_xlsx(_generation()))
    # Both values are joined in the Value cell, and the two documents line up in Source Document.
    assert "SS304 / SS316" in flat
    assert "01_Employer Spec / 02_Process" in flat
    assert "pump body SS304 / material SS316" in flat


def test_render_xlsx_vtf_fills_token() -> None:
    flat = _flat(render_xlsx(_generation()))
    assert "VTF" in flat  # defaulted token for a vtf field with no value
    assert "By Vendor" in flat  # scope token carried from the field value


def test_render_xlsx_vtf_safety_net_shows_token_and_value() -> None:
    flat = _flat(render_xlsx(_generation()))
    # A vtf field that also carries a lower-precedence value shows both, with its provenance.
    assert "VTF / 200 kW" in flat
    assert "03_Equipment List" in flat
    assert "motor 200 kW" in flat


def test_render_xlsx_is_a_valid_xlsx_container() -> None:
    data = render_xlsx(_generation())
    assert data.startswith(b"PK\x03\x04")  # OOXML/ZIP magic
