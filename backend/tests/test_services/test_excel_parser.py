"""Tests for spreadsheet → Markdown extraction (.xlsx via openpyxl, .xls via xlrd)."""

import datetime
from collections.abc import Callable
from pathlib import Path

import pytest
import xlwt
from openpyxl import Workbook

from app.services.ingestion import excel_parser
from app.services.ingestion.excel_parser import extract_xls_markdown, extract_xlsx_markdown


def _xlsx_path(tmp_path: Path, build: Callable[[Workbook], None]) -> str:
    workbook = Workbook()
    build(workbook)
    path = tmp_path / "book.xlsx"
    workbook.save(str(path))
    return str(path)


def _xls_path(tmp_path: Path, build: Callable[[xlwt.Workbook], None]) -> str:
    workbook = xlwt.Workbook()
    build(workbook)
    path = tmp_path / "book.xls"
    workbook.save(str(path))
    return str(path)


def test_xlsx_multi_sheet_with_types(tmp_path: Path) -> None:
    def build(workbook: Workbook) -> None:
        data = workbook.active
        data.title = "Data"
        data.append(["Tag", "Qty", "When"])
        data.append(["P-101", 2, datetime.datetime(2024, 1, 15)])
        data.append([None, 3.5, None])
        notes = workbook.create_sheet("Notes")
        notes.append(["a|b"])

    markdown, sheet_names = extract_xlsx_markdown(_xlsx_path(tmp_path, build))

    assert sheet_names == ["Data", "Notes"]
    assert "## Data" in markdown
    assert "## Notes" in markdown
    assert "| Tag | Qty | When |" in markdown
    assert "2024-01-15T00:00:00" in markdown  # datetime → ISO 8601
    assert "3.5" in markdown  # float preserved
    assert "a\\|b" in markdown  # pipe escaped


def test_xlsx_empty_sheet_marker(tmp_path: Path) -> None:
    markdown, sheet_names = extract_xlsx_markdown(
        _xlsx_path(tmp_path, lambda wb: setattr(wb.active, "title", "Empty"))
    )
    assert sheet_names == ["Empty"]
    assert "## Empty" in markdown
    assert "_(empty sheet)_" in markdown


def test_xlsx_truncates_large_sheet(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(excel_parser, "_MAX_ROWS_PER_SHEET", 2)
    monkeypatch.setattr(excel_parser, "_MAX_COLS_PER_SHEET", 2)

    def build(workbook: Workbook) -> None:
        sheet = workbook.active
        sheet.title = "Big"
        for row in range(4):
            sheet.append([f"r{row}c{col}" for col in range(4)])

    markdown, _ = extract_xlsx_markdown(_xlsx_path(tmp_path, build))
    assert "> _Truncated: showing first 2 of 4 rows; columns past 2 omitted._" in markdown


def test_xlsx_drops_empty_rows_and_columns(tmp_path: Path) -> None:
    def build(workbook: Workbook) -> None:
        sheet = workbook.active
        sheet.title = "Sparse"
        # Data in columns A and C only; column B and row 3 are entirely empty.
        sheet["A1"] = "Tag"
        sheet["C1"] = "Capacity"
        sheet["A2"] = "P-101"
        sheet["C2"] = 860
        sheet["A4"] = "P-202"
        sheet["C4"] = 900

    markdown, _ = extract_xlsx_markdown(_xlsx_path(tmp_path, build))

    # The empty column B and the blank row 3 are gone — no empty filler cells remain.
    assert "| Tag | Capacity |" in markdown
    assert "| P-101 | 860 |" in markdown
    assert "| P-202 | 900 |" in markdown
    assert "| |" not in markdown


def test_xls_extracts_markdown_with_dates(tmp_path: Path) -> None:
    def build(workbook: xlwt.Workbook) -> None:
        sheet = workbook.add_sheet("Sheet1")
        sheet.write(0, 0, "Tag")
        sheet.write(0, 1, "When")
        sheet.write(1, 0, "P-101")
        date_style = xlwt.easyxf(num_format_str="YYYY-MM-DD")
        sheet.write(1, 1, datetime.datetime(2024, 1, 15), date_style)

    markdown, sheet_names = extract_xls_markdown(_xls_path(tmp_path, build))

    assert sheet_names == ["Sheet1"]
    assert "## Sheet1" in markdown
    assert "| Tag | When |" in markdown
    assert "P-101" in markdown
    assert "2024-01-15" in markdown  # date cell → ISO
