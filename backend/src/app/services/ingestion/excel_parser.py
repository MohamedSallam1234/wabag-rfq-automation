"""Spreadsheet → Markdown extraction for ``.xlsx`` (openpyxl) and ``.xls`` (xlrd).

A workbook becomes a single Markdown document: each sheet is emitted in workbook order as a
``## {sheet_name}`` heading followed by a GitHub-flavored Markdown table of its data (the first
remaining row is treated as the table header). Opening the workbook is also the integrity check —
a corrupt file, or one whose real type does not match its extension (e.g. a ``.docx`` renamed
``.xlsx``), fails to load and is rejected upstream.

Fully-empty rows and columns are dropped before rendering: Excel "used ranges" routinely extend
far past the real data (a stray formatted cell stretches the sheet), and rendering that whole
rectangle verbatim produces huge runs of empty ``| |`` cells that waste downstream LLM tokens.

Very large sheets are truncated (see :data:`_MAX_ROWS_PER_SHEET` / :data:`_MAX_COLS_PER_SHEET`)
with a marker; the stored-artifact size cap remains the hard backstop.
"""

from collections.abc import Iterable, Iterator, Sequence

import xlrd
from openpyxl import load_workbook

from app.services.ingestion.markdown_table import gfm_table

#: Per-sheet caps. Beyond these the table is truncated and a marker is appended.
_MAX_ROWS_PER_SHEET = 2000
_MAX_COLS_PER_SHEET = 64
_EMPTY_SHEET_MARKER = "_(empty sheet)_"


def _is_blank(value: object) -> bool:
    """Report whether a cell value is empty (``None`` or a whitespace-only string)."""
    return value is None or (isinstance(value, str) and value.strip() == "")


def _collect_rows(
    row_iter: Iterable[Sequence[object]], max_rows: int, max_cols: int
) -> tuple[list[list[object]], bool, bool]:
    """Drain ``row_iter``, keeping at most ``max_rows`` rows of at most ``max_cols`` columns.

    Returns the kept rows (each clipped to ``max_cols``), whether any rows past ``max_rows`` were
    dropped, and whether any *non-blank* data was dropped past the column cap. Iteration stops at
    the first row beyond ``max_rows`` so cost stays bounded by the caps rather than the sheet's
    full (possibly abusive) used range — that one peek row is enough to flag the truncation. Empty
    rows/columns are not removed here — that compaction happens in :func:`_compact` once the block
    is known.
    """
    kept: list[list[object]] = []
    rows_truncated = False
    cols_truncated = False
    for row in row_iter:
        if len(row) > max_cols and any(not _is_blank(cell) for cell in row[max_cols:]):
            cols_truncated = True
        if len(kept) < max_rows:
            kept.append(list(row[:max_cols]))
        else:
            rows_truncated = True
            break
    return kept, rows_truncated, cols_truncated


def _compact(rows: list[list[object]]) -> list[list[object]]:
    """Drop fully-blank rows and fully-blank columns so phantom cells don't bloat the table.

    A column blank in *every* row is removed at the same index across all rows (so header/row
    alignment is preserved); a row blank in every column is removed. Returns ``[]`` if nothing
    non-blank remains.
    """
    rows = [row for row in rows if any(not _is_blank(cell) for cell in row)]
    if not rows:
        return []
    width = max(len(row) for row in rows)
    padded = [row + [None] * (width - len(row)) for row in rows]
    keep_cols = [c for c in range(width) if any(not _is_blank(row[c]) for row in padded)]
    if not keep_cols:
        return []
    return [[row[c] for c in keep_cols] for row in padded]


def _render_sheet(
    name: str, kept: list[list[object]], rows_truncated: bool, cols_truncated: bool
) -> str:
    """Render one sheet as a ``## {name}`` heading plus a compacted GFM table (or empty marker)."""
    heading = f"## {name}"
    compact = _compact(kept)
    if not compact:
        return f"{heading}\n\n{_EMPTY_SHEET_MARKER}"
    table = gfm_table(compact)
    notes: list[str] = []
    if rows_truncated:
        notes.append(f"showing first {len(kept)} rows")
    if cols_truncated:
        notes.append(f"columns past {_MAX_COLS_PER_SHEET} omitted")
    if notes:
        return f"{heading}\n\n{table}\n\n> _Truncated: {'; '.join(notes)}._"
    return f"{heading}\n\n{table}"


def extract_xlsx_markdown(path: str) -> tuple[str, list[str]]:
    """Convert an ``.xlsx`` workbook to Markdown; return ``(markdown, sheet_names)``.

    Uses ``read_only=True, data_only=True`` so formula cells yield their last cached value
    (``None`` if the file was never opened in Excel) rather than the formula text.

    Raises:
        Exception: If the file is not a readable ``.xlsx`` workbook.
    """
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
        blocks = []
        for name in sheet_names:
            worksheet = workbook[name]
            kept, rows_truncated, cols_truncated = _collect_rows(
                worksheet.iter_rows(values_only=True), _MAX_ROWS_PER_SHEET, _MAX_COLS_PER_SHEET
            )
            blocks.append(_render_sheet(name, kept, rows_truncated, cols_truncated))
    finally:
        workbook.close()
    return "\n\n".join(blocks), sheet_names


def _xls_cell_value(book: xlrd.book.Book, sheet: xlrd.sheet.Sheet, row: int, col: int) -> object:
    """Return a single ``.xls`` cell's value, converting serial dates to ``datetime``."""
    cell = sheet.cell(row, col)
    if cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate.xldate_as_datetime(cell.value, book.datemode)
    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return None
    return cell.value


def _xls_rows(book: xlrd.book.Book, sheet: xlrd.sheet.Sheet) -> Iterator[list[object]]:
    """Yield each row of an ``.xls`` sheet as a list of converted cell values."""
    for row in range(sheet.nrows):
        yield [_xls_cell_value(book, sheet, row, col) for col in range(sheet.ncols)]


def extract_xls_markdown(path: str) -> tuple[str, list[str]]:
    """Convert a legacy ``.xls`` workbook to Markdown; return ``(markdown, sheet_names)``.

    Raises:
        Exception: If the file is not a readable ``.xls`` workbook (e.g. an OLE2 file that is
            a ``.doc``, not a spreadsheet); the caller maps this to a 422.
    """
    book = xlrd.open_workbook(path)
    try:
        sheet_names = list(book.sheet_names())
        blocks = []
        for name in sheet_names:
            sheet = book.sheet_by_name(name)
            kept, rows_truncated, cols_truncated = _collect_rows(
                _xls_rows(book, sheet), _MAX_ROWS_PER_SHEET, _MAX_COLS_PER_SHEET
            )
            blocks.append(_render_sheet(name, kept, rows_truncated, cols_truncated))
    finally:
        book.release_resources()
    return "\n\n".join(blocks), sheet_names
